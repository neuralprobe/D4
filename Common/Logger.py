import threading
from datetime import datetime
import os
import pandas as pd
import pytz
from openpyxl import load_workbook


class Logger:
    _file_locks = {}  # 공유 파일 잠금 관리
    _file_handles = {}  # 공유 파일 핸들 관리
    _lock = threading.Lock()  # 클래스 전체에 대한 잠금
    _loggers = []

    def __init__(self, file_name):
        file_name = os.path.join(os.environ.get('D4'),'Results/'+file_name)
        with Logger._lock:
            # 파일 이름별로 단일 핸들과 잠금 생성
            if file_name not in Logger._file_locks:
                Logger._file_locks[file_name] = threading.Lock()
                Logger._file_handles[file_name] = open(file_name, "a")
        self.file_name = file_name
        self.initiated = False
        Logger._loggers.append(self)

    def _write_to_file(self, message):
        # 파일에 안전하게 쓰기
        with Logger._file_locks[self.file_name]:
            handle = Logger._file_handles[self.file_name]
            handle.write(message + "\n")
            handle.flush()  # 즉시 디스크에 기록

    def log(self, *args, **kwargs):
        # 현재 시간과 함께 메시지 작성
        # message = f"{datetime.now()} - {' '.join(map(str, args))}"
        message = f"{' '.join(map(str, args))}"
        print(message)  # 콘솔 출력
        self._write_to_file(message)  # 파일 저장

    def __call__(self, *args, **kwargs):
        # log 메서드와 동일한 동작
        self.log(*args, **kwargs)

    def close(self):
        # 현재 파일 핸들 닫기
        with Logger._lock:
            if self.file_name in Logger._file_handles:
                Logger._file_handles[self.file_name].close()
                del Logger._file_handles[self.file_name]
                del Logger._file_locks[self.file_name]

    @staticmethod
    def close_all():
        for logger in Logger._loggers:
            logger.close()

def search_and_export_to_excel(filename, start, end,):
    folder_path = os.path.join(os.environ.get('D4'), 'Results')
    output_excel = filename + "_summary" + f"_{start}_{end}_{datetime.now(pytz.timezone('America/New_York')).strftime('%Y-%m-%d %H-%M-%S')}.xlsx"
    output_excel = os.path.join(folder_path,output_excel)
    writer = pd.ExcelWriter(output_excel, engine='openpyxl')
    # 폴더 내 파일 목록 검색
    for file in os.listdir(folder_path):
        if filename in file and start in file and end in file and file.endswith(".csv"):
            file_path = os.path.join(folder_path, file)
            try:
                data = pd.read_csv(file_path)
            except Exception as e:
                print(f"파일 읽기 실패: {file_path}, 에러: {e}")
                continue

            # 파일명에 따라 시트 이름 결정
            if "account" in file:
                sheet_name = "account"
            elif "order" in file:
                sheet_name = "order"
            elif "prophecy" in file:
                sheet_name = "prophecy"
            elif "trader" in file:
                sheet_name = "trader"
            else:
                sheet_name = os.path.splitext(file)[0]

            # 데이터 추가
            data.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"파일 {file}을(를) {sheet_name} 시트에 추가했습니다.")

    if len(writer.sheets) == 0:
        # 기본 시트를 생성
        empty_df = pd.DataFrame({"No Data": ["This sheet is empty"]})
        empty_df.to_excel(writer, sheet_name="Default", index=False)

    # Excel 파일 저장 및 기본 Sheet 제거
    writer.close()

    # 기본 Sheet 제거
    workbook = load_workbook(output_excel)
    if 'Sheet' in workbook.sheetnames:  # 기본 시트가 존재할 경우 삭제
        workbook.remove(workbook['Sheet'])
    workbook.save(output_excel)
    print(f"Excel 파일이 생성되었습니다: {output_excel}")

# 사용 예제
if __name__ == "__main__":
    # 로거 생성
    logger1 = Logger("example.log")
    logger2 = Logger("example.log")  # 동일한 파일 이름 사용

    logger1.log("This is a test message from logger1.")
    logger2.log("This is a test message from logger2.")

    # print 스타일 사용
    logger1("Logger1 can be used like print.")

    # 로거 닫기
    logger1.close()
    logger2.close()
